import logging
import os

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from typing import Optional

from dbs.entities import Data, Info, db
from domain.enums import SheetNames, DataSheetIndexes, InfoSheetIndexes
from domain.logger import logged


class DatabaseSession:

    def __init__(self, database: str, session: Optional[Session] = None):
        session_maker = sessionmaker()
        session_maker.configure(bind=database)
        self.session = session_maker() if session is None else session

    def __enter__(self) -> Session:
        return self.session

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is not None:
            self.session.rollback()
            self.session.close()
            raise exc_val

        self.session.commit()
        self.session.close()


class DatabaseInitializer:

    def __init__(self, database: str, file: str, logger: logging.Logger):
        engine = create_engine(database)
        self.recreate_tables(engine)

        session_maker = sessionmaker()
        session_maker.configure(bind=engine)
        self.session = session_maker()

        self.logger = logger
        self.workbook = load_workbook(file, read_only=True, data_only=True)

    @staticmethod
    def recreate_tables(engine):
        tables = [t.__table__ for t in [
            Data,
            Info
        ]]
        db.metadata.drop_all(bind=engine, tables=tables)
        db.metadata.create_all(bind=engine, tables=tables)

    @logged
    def init_data(self):
        data = []
        for row in self.workbook[SheetNames.data.value].iter_rows(min_row=2, values_only=True):
            if row[DataSheetIndexes.id] is None:
                break
            data.append(Data(
                id=row[DataSheetIndexes.id],
                edu_index=row[DataSheetIndexes.edu_index],
                science_index=row[DataSheetIndexes.science_index],
                inter_index=row[DataSheetIndexes.inter_index],
                fin_index=row[DataSheetIndexes.fin_index],
                wage=row[DataSheetIndexes.wage],
                add=row[DataSheetIndexes.add],
                exam_1=row[DataSheetIndexes.exam_1],
                exam_2=row[DataSheetIndexes.exam_2],
                exam_3=row[DataSheetIndexes.exam_3],
                avg_exam_1=row[DataSheetIndexes.avg_exam_1],
                stud_num_1=row[DataSheetIndexes.stud_num_1],
                stud_num_2=row[DataSheetIndexes.stud_num_2],
                stud_num_3=row[DataSheetIndexes.stud_num_3],
                w_1=row[DataSheetIndexes.w_1],
                w_2=row[DataSheetIndexes.w_2],
                w_3=row[DataSheetIndexes.w_3],
                w_4=row[DataSheetIndexes.w_4],
                w_5=row[DataSheetIndexes.w_5],
                asp_num=row[DataSheetIndexes.asp_num],
                w_6=row[DataSheetIndexes.w_6],
                w_7=row[DataSheetIndexes.w_7],
                citation_1=row[DataSheetIndexes.citation_1],
                citation_2=row[DataSheetIndexes.citation_2],
                citation_3=row[DataSheetIndexes.citation_3],
                pub_num_1=row[DataSheetIndexes.pub_num_1],
                pub_num_2=row[DataSheetIndexes.pub_num_2],
                pub_num_3=row[DataSheetIndexes.pub_num_3],
                research_num=row[DataSheetIndexes.research_num],
                w_8=row[DataSheetIndexes.w_8],
                w_9=row[DataSheetIndexes.w_9],
                research_incom=row[DataSheetIndexes.research_incom],
                license_num=row[DataSheetIndexes.license_num],
                w_10=row[DataSheetIndexes.w_10],
                w_11=row[DataSheetIndexes.w_11],
                w_12=row[DataSheetIndexes.w_12],
                journal_num=row[DataSheetIndexes.journal_num],
                grants_num=row[DataSheetIndexes.grants_num],
                w_13=row[DataSheetIndexes.w_13],
                w_14=row[DataSheetIndexes.w_14],
                w_15=row[DataSheetIndexes.w_15],
                w_16=row[DataSheetIndexes.w_16],
                w_17=row[DataSheetIndexes.w_17],
                w_18=row[DataSheetIndexes.w_18],
                stud_num_4=row[DataSheetIndexes.stud_num_4],
                w_19=row[DataSheetIndexes.w_19],
                prof_num=row[DataSheetIndexes.prof_num],
                w_20=row[DataSheetIndexes.w_20],
                w_21=row[DataSheetIndexes.w_21],
                money_1=row[DataSheetIndexes.money_1],
                money_2=row[DataSheetIndexes.money_2],
                income_per_npr=row[DataSheetIndexes.income_per_npr],
                w_22=row[DataSheetIndexes.w_22],
                avg_wage_1=row[DataSheetIndexes.avg_wage_1],
                total_income=row[DataSheetIndexes.total_income],
                total_square_1=row[DataSheetIndexes.total_square_1],
                square_1=row[DataSheetIndexes.square_1],
                square_2=row[DataSheetIndexes.square_2],
                square_3=row[DataSheetIndexes.square_3],
                square_4=row[DataSheetIndexes.square_4],
                computer_number=row[DataSheetIndexes.computer_number],
                w_23=row[DataSheetIndexes.w_23],
                books=row[DataSheetIndexes.books],
                w_24=row[DataSheetIndexes.w_24],
                w_25=row[DataSheetIndexes.w_25],
                w_26=row[DataSheetIndexes.w_26],
                npr_num=row[DataSheetIndexes.npr_num],
                w_27=row[DataSheetIndexes.w_27],
                total_stud_num_1=row[DataSheetIndexes.total_stud_num_1],
                stud_num_5=row[DataSheetIndexes.stud_num_5],
                stud_num_6=row[DataSheetIndexes.stud_num_6],
                stud_num_7=row[DataSheetIndexes.stud_num_7],
                avg_exam_2=row[DataSheetIndexes.avg_exam_2],
                w_28=row[DataSheetIndexes.w_28],
                w_29=row[DataSheetIndexes.w_29],
                w_30=row[DataSheetIndexes.w_30],
                stud_num_8=row[DataSheetIndexes.stud_num_8],
                stud_num_9=row[DataSheetIndexes.stud_num_9],
                company_num_1=row[DataSheetIndexes.company_num_1],
                company_num_2=row[DataSheetIndexes.company_num_2],
                money_3=row[DataSheetIndexes.money_3],
                money_4=row[DataSheetIndexes.money_4],
                total_pub_num=row[DataSheetIndexes.total_pub_num],
                bi_num=row[DataSheetIndexes.bi_num],
                tp_num=row[DataSheetIndexes.tp_num],
                ccus_num=row[DataSheetIndexes.ccus_num],
                sb_num=row[DataSheetIndexes.sb_num],
                total_asp_num=row[DataSheetIndexes.total_asp_num],
                w_31=row[DataSheetIndexes.w_31],
                total_doc_num=row[DataSheetIndexes.total_doc_num],
                diss_sov_num=row[DataSheetIndexes.diss_sov_num],
                total_work_num=row[DataSheetIndexes.total_work_num],
                total_pps_num=row[DataSheetIndexes.total_pps_num],
                total_sci_num=row[DataSheetIndexes.total_sci_num],
                w_32=row[DataSheetIndexes.w_32],
                w_33=row[DataSheetIndexes.w_33],
                w_34=row[DataSheetIndexes.w_34],
                w_35=row[DataSheetIndexes.w_35],
                avg_wage_2=row[DataSheetIndexes.avg_wage_2],
                avg_wage_3=row[DataSheetIndexes.avg_wage_3],
                for_stud_num=row[DataSheetIndexes.for_stud_num],
                w_36=row[DataSheetIndexes.w_36],
                total_prog_num=row[DataSheetIndexes.total_prog_num],
                total_stud_num_2=row[DataSheetIndexes.total_stud_num_2],
                total_for_asp_num=row[DataSheetIndexes.total_for_asp_num],
                citation_4=row[DataSheetIndexes.citation_4],
                for_income_1=row[DataSheetIndexes.for_income_1],
                for_income_2=row[DataSheetIndexes.for_income_2],
                results_num=row[DataSheetIndexes.results_num],
                total_square_2=row[DataSheetIndexes.total_square_2],
                lab_square=row[DataSheetIndexes.lab_square],
                research_square=row[DataSheetIndexes.research_square],
                living_square=row[DataSheetIndexes.living_square],
                sport_square=row[DataSheetIndexes.sport_square],
                w_37=row[DataSheetIndexes.w_37],
                comp_num=row[DataSheetIndexes.comp_num],
                w_38=row[DataSheetIndexes.w_38],
                library=row[DataSheetIndexes.library],
                income=row[DataSheetIndexes.income],
                non_budget_income=row[DataSheetIndexes.non_budget_income],
                w_39=row[DataSheetIndexes.w_39],
                w_40=row[DataSheetIndexes.w_40],
                w_41=row[DataSheetIndexes.w_41],
                w_42=row[DataSheetIndexes.w_42],
                w_43=row[DataSheetIndexes.w_43],
                w_44=row[DataSheetIndexes.w_44],
                w_45=row[DataSheetIndexes.w_45]
            ))
        self.session.bulk_save_objects(data)

    @logged
    def init_info(self):
        info = []
        for row in self.workbook[SheetNames.info.value].iter_rows(min_row=2, values_only=True):
            if row == ():
                break
            info.append(Info(
                id=row[InfoSheetIndexes.id],
                var_name=row[InfoSheetIndexes.var_name],
                var_description=row[InfoSheetIndexes.var_description],
            ))
        self.session.bulk_save_objects(info)

    @logged
    def init_db(self):
        self.init_data()
        self.init_info()

        self.session.commit()
